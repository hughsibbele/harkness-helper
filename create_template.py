"""
Creates the Harkness Helper spreadsheet template as an .xlsx file.
Tabs: Settings, Discussions, Students, Transcripts, SpeakerMap, StudentReports, Prompts, Courses
Headers match exactly what initializeSheetHeaders() in Sheets.gs produces.
Run: python3 create_template.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# --- Styles ---
header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4285F4", end_color="4285F4", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
body_font = Font(name="Arial", size=10)
body_align = Alignment(vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

def style_header(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}1"


# ============================================================
# Tab 1: Settings (key-value config)
# ============================================================
ws_settings = wb.active
ws_settings.title = "Settings"

ws_settings.append(["setting_key", "setting_value"])
style_header(ws_settings, 2)

settings_defaults = [
    ["mode", "group"],
    ["distribute_email", "true"],
    ["distribute_canvas", "false"],
    ["grade_scale", "0-100"],
    ["teacher_email", ""],
    ["teacher_name", ""],
    ["email_subject_template", "Harkness Discussion Report - {date}"],
    ["gemini_model", "gemini-2.0-flash"],
    ["elevenlabs_model", "scribe_v2"],
    ["canvas_course_id", ""],
    ["canvas_base_url", ""],
    ["canvas_item_type", "assignment"],
]

for row in settings_defaults:
    ws_settings.append(row)

ws_settings.column_dimensions["A"].width = 30
ws_settings.column_dimensions["B"].width = 50


# ============================================================
# Tab 2: Discussions
# ============================================================
ws_disc = wb.create_sheet("Discussions")

disc_headers = [
    "status", "next_step", "date", "section", "course",
    "grade", "approved", "group_feedback",
    "canvas_assignment_id", "canvas_item_type",
    "discussion_id", "audio_file_id", "error_message",
    "created_at", "updated_at",
]
ws_disc.append(disc_headers)
style_header(ws_disc, len(disc_headers))

disc_widths = [12, 30, 12, 12, 14, 8, 10, 40, 18, 14, 18, 18, 30, 18, 18]
for i, w in enumerate(disc_widths, 1):
    ws_disc.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# Tab 3: Students
# ============================================================
ws_stu = wb.create_sheet("Students")

stu_headers = ["name", "email", "section", "course", "canvas_user_id", "student_id"]
ws_stu.append(stu_headers)
style_header(ws_stu, len(stu_headers))

stu_widths = [22, 28, 14, 14, 14, 18]
for i, w in enumerate(stu_widths, 1):
    ws_stu.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# Tab 4: Transcripts
# ============================================================
ws_trans = wb.create_sheet("Transcripts")

trans_headers = [
    "discussion_id", "raw_transcript", "speaker_map", "named_transcript",
    "created_at", "updated_at",
]
ws_trans.append(trans_headers)
style_header(ws_trans, len(trans_headers))

trans_widths = [18, 60, 30, 60, 18, 18]
for i, w in enumerate(trans_widths, 1):
    ws_trans.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# Tab 5: SpeakerMap
# ============================================================
ws_sm = wb.create_sheet("SpeakerMap")

sm_headers = ["discussion_id", "speaker_label", "suggested_name", "student_name", "confirmed"]
ws_sm.append(sm_headers)
style_header(ws_sm, len(sm_headers))

sm_widths = [18, 14, 20, 20, 10]
for i, w in enumerate(sm_widths, 1):
    ws_sm.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# Tab 6: StudentReports
# ============================================================
ws_rep = wb.create_sheet("StudentReports")

rep_headers = [
    "student_name", "grade", "approved", "sent", "feedback",
    "discussion_id", "transcript_contributions", "participation_summary",
    "student_id", "report_id", "created_at", "updated_at",
]
ws_rep.append(rep_headers)
style_header(ws_rep, len(rep_headers))

rep_widths = [20, 8, 10, 8, 40, 18, 40, 30, 18, 18, 18, 18]
for i, w in enumerate(rep_widths, 1):
    ws_rep.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# Tab 7: Prompts
# ============================================================
ws_prompts = wb.create_sheet("Prompts")

ws_prompts.append(["prompt_name", "prompt_text"])
style_header(ws_prompts, 2)

prompts = [
    [
        "SPEAKER_IDENTIFICATION",
        """You are analyzing the beginning of a classroom Harkness discussion recording.

Students typically introduce themselves at the start. Listen for patterns like:
- "Hi, I'm [name]"
- "My name is [name]"
- "This is [name]"
- "[Name] here"
- Other natural introductions

Analyze this transcript excerpt and identify which speaker label corresponds to which student name.

IMPORTANT RULES:
1. Only include speakers you can confidently identify from explicit introductions
2. If a speaker cannot be identified, map them to "?" instead
3. Be case-sensitive with names as they appear
4. The teacher may also speak — if identified, include them as "Teacher"

Return ONLY a valid JSON object mapping speaker labels to names.
Example format: {"Speaker 0": "Maria", "Speaker 1": "James", "Speaker 2": "Teacher", "Speaker 3": "?"}

Transcript excerpt:
{transcript}

JSON mapping:""",
    ],
    [
        "GROUP_FEEDBACK",
        """You are a high school teacher analyzing a Harkness discussion. You will produce exactly two paragraphs.

**PARAGRAPH 1 — Discussion Summary** (Neutral Voice)
Write in a neutral, objective, third-person voice. Provide a detailed summary of the discussion's main topics and flow. Identify 2-3 "defining moments" — key turning points, breakthrough ideas, or significant challenges that shaped the conversation.

**PARAGRAPH 2 — Evaluative Comment** (Teacher Voice)
Write in the teacher's voice, directed at the class ("you" plural, "I" for the teacher). The tone must be direct, informal, supportive, and clear. Follow this mandatory "Critique Sandwich" structure:

1. **The Grade**: State the grade clearly and colloquially in the first sentence. (e.g., "This was a strong discussion, earning a solid 8.5 out of 10.", "This was a decent but not great start... 7/10.")
2. **The Good**: Highlight 2-3 specific positive achievements. Credit specific students by name, linking them to their idea or contribution.
3. **The Gap**: Identify the primary weakness or area for growth.
4. **The Next Step**: Conclude with a single, clear, actionable goal for the next discussion.

**Tone alignment with grade:**
- High grade (9-10): Frame positives as "excellent" or "deep"; the gap is a "final step" to the next level.
- Medium grade (7-8.5): Balanced ("solid," "decent start") with a more significant gap to work on.
- Lower grade (below 7): Honest but encouraging; clear gap with concrete next steps.

**Important:**
- If the teacher gave oral feedback during the discussion (often near the end — look for phrases like "my evaluation," "my feedback," or the teacher summarizing), align your evaluation with their points.
- Credit specific students by name for notable contributions.
- If the teacher intervened to guide the discussion, acknowledge this (e.g., "I had to provide the key synthesizing question").

Grade: {grade}

Transcript:
{transcript}

Write the two paragraphs now (summary paragraph first, then evaluative comment):""",
    ],
    [
        "INDIVIDUAL_FEEDBACK",
        """You are a high school teacher providing personalized feedback to {student_name} about their Harkness discussion participation. You will produce exactly two paragraphs.

**PARAGRAPH 1 — Contribution Summary** (Neutral Voice)
Write in a neutral, objective voice. Summarize what {student_name} contributed to the discussion — their main points, arguments, and how they engaged with other students' ideas. Note specific moments where they advanced or redirected the conversation.

**PARAGRAPH 2 — Evaluative Comment** (Teacher Voice)
Write in the teacher's voice, directed at the student ("you"). The tone must be direct, informal, supportive, and clear. Follow this "Critique Sandwich" structure:

1. **The Grade**: State the grade clearly in the first sentence.
2. **The Good**: Highlight 2-3 specific strengths from their participation, referencing actual points they made.
3. **The Gap**: Identify their primary area for growth as a discussion participant.
4. **The Next Step**: Conclude with a single, actionable goal for their next discussion.

**Tone alignment with grade:**
- High grade (9-10): "Excellent" contributions; the gap is a stretch goal.
- Medium grade (7-8.5): "Solid" participation with clear room to grow.
- Lower grade (below 7): Encouraging but honest about what's missing.

**Important:**
- If the teacher gave oral feedback during the discussion (often near the end — look for phrases like "my evaluation," "my feedback," or the teacher summarizing), align your evaluation with their points.

Grade: {grade}

{student_name}'s contributions:
{contributions}

Full discussion transcript (for context):
{transcript}

Write the two paragraphs now (contribution summary first, then evaluative comment for {student_name}):""",
    ],
]

for row in prompts:
    ws_prompts.append(row)

ws_prompts.column_dimensions["A"].width = 28
ws_prompts.column_dimensions["B"].width = 100

# Wrap text in prompt_text column
for row in ws_prompts.iter_rows(min_row=2, min_col=2, max_col=2):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")


# ============================================================
# Tab 8: Courses (multi-course Canvas config)
# ============================================================
ws_courses = wb.create_sheet("Courses")

courses_headers = ["course_name", "canvas_course_id", "canvas_base_url", "canvas_item_type"]
ws_courses.append(courses_headers)
style_header(ws_courses, len(courses_headers))

courses_widths = [24, 18, 36, 14]
for i, w in enumerate(courses_widths, 1):
    ws_courses.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# Save
# ============================================================
output_path = "Harkness_Helper_Template.xlsx"
wb.save(output_path)
print(f"Template saved to: {output_path}")
